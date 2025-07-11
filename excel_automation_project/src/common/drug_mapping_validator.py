import sys
from pathlib import Path

# Get current script's directory
current_dir = Path(__file__).resolve().parent
# Go up one level to project root
project_root = current_dir.parent
# Add project root to Python path
sys.path.append(str(project_root))

import pandas as pd
from openpyxl import load_workbook
from common.config_handler import config
from common.data_manager import DataManager
from common.error_handler import error_manager, ExpectedCustomError

class DrugMappingValidator:
    def __init__(self, data_manager, database_date_validation):
        try:
            self.database_date_validation = database_date_validation
            error_manager.log_info("Initializing DrugMappingValidator...")
            
            self._check_required_files()

            self.dm = data_manager
            self.optoviks = {}
            self.validation_errors = {
                "column_count": [], 
                "mandatory_columns": [],
                "sheet_names": []
            }

            self.optoviks_sh = pd.ExcelFile(config.source_path_for_optiviks)
            self.dictionaries = pd.ExcelFile(config.source_path_for_dictionary)
            self.budg_df = pd.read_excel(config.path_for_budget_difference)
            self.drug_groups_df = pd.read_excel(config.source_path_drug_groups)

            error_manager.log_info("  Validator initialized successfully")

        except ExpectedCustomError:
            raise
        except Exception:
            raise


    def _check_required_files(self):
        required_files = {
            "Optoviks": config.source_path_for_optiviks,
            "Drug Dictionary": config.source_path_for_dictionary,
            "Budget Difference": config.path_for_budget_difference,
            "Drug Groups": config.source_path_drug_groups
        }

        missing_files = []
        for name, path in required_files.items():
            if not Path(path).exists():
                missing_files.append((name, path))

        if missing_files:
            error_lines = ["\n\n" + "=" * 50 + "\nFILES ARE MISSING\n" + "=" * 50 + "\n"]
            error_lines.append("\n❌ The following required files are missing:\n")
            for name, path in missing_files:
                error_lines.append(f"• {name} — Missing file: {Path(path).name}\n  📁 Expected at: {path}\n")

            error_lines.append("\n📌 HOW TO FIX:\n")
            error_lines.append("1. Make sure each file listed above exists at its expected path.\n")
            error_lines.append("2. If missing, export or copy the correct version into the folder.\n")
            error_lines.append("3. Make sure none of the files are open in Excel or locked by another program.\n")
            error_lines.append("4. If the file paths are wrong in config, update them accordingly.\n")
            error_lines.append("\n⚠️ The script cannot continue unless all required files are present in expected location.")

            raise ExpectedCustomError("".join(error_lines))
        
    def load_optoviks(self):
        """Load source data and drug name dictionaries with cleaning"""
        try:
            error_manager.log_info("Loading optoviks data...")
            for sheet in self.optoviks_sh.sheet_names:
                optovik_df = pd.read_excel(self.optoviks_sh, sheet_name=sheet)
                self._validate_optoviks(sheet, optovik_df, 8, self.database_date_validation)

            if any(self.validation_errors.values()):
                error_msg_optovik = ["\n\n" + "="*50 + "\nDATA VALIDATION FAILED\n" + "="*50]

                error_footer = "\n\n⚠️ The script will not continue until the errors are corrected."

                if self.validation_errors['column_count']:
                    how_to_fix_column_count = (
                        "\n\n🛠 HOW TO FIX:"
                        "\n1. Open the sheets in your optoviklar file."
                        "\n2. Check the number of columns — you should have exactly 8 columns:"
                        "\n   - [Optional] Column headers must match the expected format (if applicable)."
                        "\n3. Remove any extra or unintended columns (e.g., empty, duplicate, or helper columns)."
                        "\n4. Make sure there are no hidden columns accidentally added."
                        "\n5. Save the corrected file."
                        "\n__________________________________________________"
                    )
                    body = "".join(self.validation_errors['column_count'])
                    column_count_error = body + how_to_fix_column_count
                    error_msg_optovik.append(column_count_error)

                if self.validation_errors['mandatory_columns']:
                    if self.database_date_validation:
                        how_to_fix_mandatory_columns = (
                            "\n\n🛠 HOW TO FIX:"
                            "\n1. Open the sheets in your optoviklar file."
                            "\n2. Locate the rows where the specified columns ('Клиент', 'Препарат' or 'Дата' in sheets) are empty or contain invalid values."
                            "\n3. Fill in the missing values with the correct data, or remove the rows if they are not needed."
                            "\n   - If date column has nan values Please enter the date in valid date format to all date columns: 'dd-mm-yyyy'"
                            "\n4. Make sure there are no unintended empty rows at the bottom of the sheet."
                            "\n5. Save the Excel file."
                            "\n__________________________________________________"
                        )
                    else:
                        how_to_fix_mandatory_columns = (
                            "\n\n🛠 HOW TO FIX:"
                            "\n1. Open the sheets in your optoviklar file."
                            "\n2. Locate the rows where the specified columns ('Клиент' or 'Препарат'  in sheets) are empty or contain invalid values."
                            "\n3. Fill in the missing values with the correct data, or remove the rows if they are not needed."
                            "\n4. Make sure there are no unintended empty rows at the bottom of the sheet."
                            "\n5. Save the Excel file."
                            "\n__________________________________________________"
                        )
                    body = "".join(self.validation_errors['mandatory_columns'])
                    mandatory_columns = body + how_to_fix_mandatory_columns
                    error_msg_optovik.append(mandatory_columns)
                    
                if self.validation_errors['sheet_names']:
                    how_to_fix_sheet_names = (
                        f"\n\n📄 Available sheets in Drug Dictionary: {self.dictionaries.sheet_names}\n"
                        "\n\n🛠 HOW TO FIX:\n"
                        "1. Open the Drug Dictionary Excel file.\n"
                        "2. Make sure that **every sheet used in the Optoviks file** exists in the Drug Dictionary file.\n"
                        "3. Sheet names must match **exactly** (including case and spacing).\n"
                        "4. Check for common issues:\n"
                        "   - Capitalization differences (e.g., 'Meros' ≠ 'meros')\n"
                        "   - Extra spaces or invisible characters\n"
                        "   - Typos or alternate spellings\n"
                        "5. Rename or add the missing sheets to the Drug Dictionary file.\n"
                        "6. Save the file and re-run the script.\n\n"
                        "\n__________________________________________________"

                    )
                    body = "".join(self.validation_errors['sheet_names'])
                    sheet_names = body + how_to_fix_sheet_names
                    error_msg_optovik.append(sheet_names)

                error_msg_optovik.append(error_footer)
                error_msg_optovik = "\n".join(error_msg_optovik)
                    
                raise ExpectedCustomError(error_msg_optovik)
                
            # LOADING CUSTOMERS DF
            for sheet in self.optoviks_sh.sheet_names:
                optovik_df = pd.read_excel(self.optoviks_sh, sheet_name=sheet)
                # changing header names dynamically
                optovik_df = optovik_df.rename(columns={
                    optovik_df.columns[config.optivik_drug_col_index]: config.drugs_header_name,
                    optovik_df.columns[config.optivik_customer_col_index]: config.client_header_name,
                    optovik_df.columns[config.optivik_region_col_index]: config.region_header_name,
                    optovik_df.columns[config.optivik_territory_col_index]: config.territory_header_name,
                    optovik_df.columns[config.optivik_quantity_col_index]: config.quantity_header_name,
                    optovik_df.columns[config.optivik_price_col_index]: config.price_header_name,
                    optovik_df.columns[config.optivik_reserve_col_index]: config.reserve_header_name,
                    optovik_df.columns[config.optivik_month_year_col_index]: config.date_header_name,
                })

                # Cleaning and formatting
                optovik_df[config.territory_header_name] = optovik_df[config.territory_header_name].astype(str).str.strip()
                optovik_df[config.region_header_name] = optovik_df[config.region_header_name].astype(str).str.strip()
                optovik_df[config.drugs_header_name] = optovik_df[config.drugs_header_name].astype(str).str.strip()
                optovik_df[config.client_header_name] = optovik_df[config.client_header_name].astype(str).str.strip()
                optovik_df[config.reserve_header_name] = optovik_df[config.reserve_header_name].astype(str).str.strip()
                optovik_df[config.date_header_name] = pd.to_datetime(optovik_df[config.date_header_name], errors='coerce', format='%d.%m.%Y')

                for col in [config.quantity_header_name, config.price_header_name]:
                    optovik_df[col] = (
                        optovik_df[col]
                        .replace('', '0')
                        .astype(float)
                        .fillna(0)
                    )
                # calculate total sales
                optovik_df[config.total_sales_header_name] = optovik_df[config.price_header_name] * optovik_df[config.quantity_header_name]
                optovik_df = optovik_df.drop(config.price_header_name, axis=1)
                # store data
                self.optoviks[sheet] = optovik_df
            error_manager.log_info(f"  Loaded {len(self.optoviks_sh.sheet_names)} optovik sheets successfully")
        except Exception:
            raise

    def _validate_optoviks(self, sheet_name, df, required_columns, database_check: False):
        """Validate loaded Optoviks data for critical issues"""
        try:
            # Required column count check
            REQUIRED_COLUMN_COUNT = required_columns
            
            # 1. Column count validation
            if len(df.columns) != REQUIRED_COLUMN_COUNT:
                self.validation_errors['column_count'].append(
                    f"\n\n❌ Sheet '{sheet_name}' has {len(df.columns)} columns, but exactly {REQUIRED_COLUMN_COUNT} columns are required."
                    )

            # 2. Mandatory columns non-NaN check
            if database_check:
                mandatory_columns = [
                    config.optivik_drug_col_index,   # Preparat
                    config.optivik_customer_col_index, # Klient
                    config.optivik_month_year_col_index # date
                ]
            else:
                mandatory_columns = [
                    config.optivik_drug_col_index,   # Preparat
                    config.optivik_customer_col_index
                ]

            for ind, col in enumerate(mandatory_columns):
                if df.iloc[:, col].isna().any():
                    na_count = df.iloc[:, col].isna().sum()
                    self.validation_errors["mandatory_columns"].append(
                        f"\n\n❌ Sheet '{sheet_name}': '{config.drugs_header_name if ind == 0 else config.client_header_name if ind == 1 else config.date_header_name}' column contains {na_count} missing (NaN) values. "
                    )

            # 3 sheet name validation
            if sheet_name not in self.dictionaries.sheet_names:
                self.validation_errors["sheet_names"].append(
                    f"\n\n❌ The sheet '{sheet_name}' exists in the Optoviks file but was not found in the Drug Dictionary Excel file."
                )
        except Exception as e:
            raise RuntimeError(f"Error validating optoviks sheet '{sheet_name}': {str(e)}")

    def dublicate_client_region_validation(self):
        try:
            error_manager.log_info("Validating client region duplicates...")
            l = []
            for sheet, df in self.optoviks.items():
                df['Optoviks'] = sheet
                l.append(df)
            main_df = pd.concat(l)
            
            # Step 1: Group by raw 'client', count unique regions and territories
            grouped = main_df.groupby(config.client_header_name)[[config.region_header_name, config.territory_header_name]].nunique()

            # Step 2: Keep only clients with more than 1 unique region or territory
            conflicting_clients = grouped[(grouped[config.region_header_name] > 1) | (grouped[config.territory_header_name] > 1)].index

            # Step 3: Filter the original DataFrame
            result = (
                main_df[main_df[config.client_header_name].isin(conflicting_clients)]
                .drop_duplicates(subset=[config.client_header_name, config.region_header_name, config.territory_header_name])
            )[['Optoviks', config.client_header_name, config.region_header_name, config.territory_header_name]].sort_values(by=config.client_header_name)
            
            if not result.empty:
                result.to_excel(config.duplicate_clients, index=False)
                message = (
                    "\n\n" + "=" * 50 + "\nDATA VALIDATION FAILED\n" + "=" * 50 + "\n"
                    "\n❌ Duplicate Clients with Conflicting Regions or Territories\n"
                    "One or more clients (drug store names) appear multiple times with **different Region or Territory** values.\n"
                    f"\n→ Please review and correct the entries in the Excel file:\n{config.duplicate_clients}\n"
                    "\n🛠 HOW TO FIX:"
                    "- Ensure each client appears with **only one Region and one Territory**\n"
                    "- Correct inconsistencies in the original input files (check spelling, spacing, casing)\n"
                    "- Re-run the script after fixing the data\n"
                    "\n🔍 Tip: Use Excel filters or conditional formatting to spot differences easily.\n"
                )
                raise ExpectedCustomError(message)
            error_manager.log_info("  Client region validation completed successfully")
        except Exception:
            raise

    def load_dictionary(self):
        try:
            error_manager.log_info("Loading drug dictionaries...")
            for sheet in self.dictionaries.sheet_names:
                # if customer not in the optiviks, don't load its dictionary
                if sheet not in self.optoviks:
                    continue
                # Load and clean dictionary
                dict_df = pd.read_excel(config.source_path_for_dictionary, sheet_name=sheet)
                # Remove completely empty rows
                dict_df = dict_df.dropna(how='all')

                # handling dynamic col position
                dict_df = dict_df.rename(columns={
                    dict_df.columns[config.dict_cust_drug_names_index]: config.cust_drug_names,
                    dict_df.columns[config.dict_std_drug_names_index]: config.std_drug_names
                })
                # Convert to string and clean
                dict_df[config.cust_drug_names] = (dict_df[config.cust_drug_names].astype(str).str.strip().replace({'nan': pd.NA, '': pd.NA}))
                dict_df[config.std_drug_names] = (dict_df[config.std_drug_names].astype(str).str.strip().replace({'nan': pd.NA, '': pd.NA}))
                # Add to dictionaries
                self.dm.add_dictionary(sheet, dict_df)
            error_manager.log_info(f"  Loaded {len(self.dm.drug_name_dict)} drug dictionaries")
        except Exception:
            raise

    def validate_dictionary(self):
        """Validate drug name mapping dictionary for data quality issues"""
        try:
            error_manager.log_info("Validating drug dictionaries...")
            errors = {
                'missing_names': {},
                'missing_standards': {},
                'duplicates': {}
            }
            for sheet, df in self.dm.drug_name_dict.items():
                # Check 1: Missing customer drug names
                missing_names = df[config.cust_drug_names].isna().sum()
                if missing_names:
                    errors['missing_names'][sheet] = missing_names
                # Check 2: Customer names without standards
                missing_std = df.loc[df[config.cust_drug_names].notna() & df[config.std_drug_names].isna()].shape[0]
                if missing_std:
                    errors['missing_standards'][sheet] = missing_std
                # Check 3: Duplicates (excluding nulls)
                valid_entries = df.dropna(subset=[config.cust_drug_names])
                dupes = valid_entries[valid_entries.duplicated(config.cust_drug_names, keep=False)]
                if not dupes.empty:
                    errors['duplicates'][sheet] = dupes[config.cust_drug_names].unique().tolist()
            # Error reporting
            if any(errors.values()):
                error_msg = ["\n\n" + "=" * 50 + "\nDRUG NAME DICTIONARY VALIDATION FAILED\n" + "=" * 50 + "\n"]
                
                
                if errors['missing_names']:
                    error_msg.append(f"❌ MISSING {config.cust_drug_names}:")
                    for sheet, count in errors['missing_names'].items():
                        error_msg.append(f"  - {sheet}: {count} empty/missing {config.cust_drug_names}")

                if errors['missing_standards']:
                    error_msg.append("❌ MISSING STANDARD NAMES:")
                    for sheet, count in errors['missing_standards'].items():
                        error_msg.append(f"  - {sheet}: {count} entries have {config.cust_drug_names} but no standard mapping")

                if errors['duplicates']:
                    error_msg.append(f"❌ DUPLICATE {config.cust_drug_names}:")
                    for sheet, names in errors['duplicates'].items():
                        error_msg.append(f"  - {sheet}: Duplicates found for - {', '.join(names)}")

                error_msg.append("\n🛠 HOW TO FIX:")
                if errors['missing_names']:
                    error_msg.append(f"  • Fill empty cells in {config.cust_drug_names} column")
                if errors['missing_standards']:
                    error_msg.append("  • Add standard names for customer entries that have drug names")
                if errors['duplicates']:
                    error_msg.append(f"  • Remove duplicate {config.cust_drug_names} or ensure consistent standard name mappings")

                error_msg.append(f"\n⚠️ Fix these issues in the dictionary file and try again.")
                
                raise ExpectedCustomError('\n'.join(error_msg))
            error_manager.log_info("  Drug dictionaries validated successfully")
        except Exception:
            raise

    def validate_unmapped_drugs(self):
        """Validate all sheets for missing drug mappings"""
        try:
            error_manager.log_info("Validating unmapped drugs...")
            missing_drugs = {}
            for sheet in self.dm.drug_name_dict.keys():                       
                try:
                    optovik_df = self.optoviks[sheet]
                    # Read and Clean dictionary drug names
                    dict_df = self.dm.drug_name_dict[sheet]
                    # Merge with cleaned values
                    merged = optovik_df.merge(
                        dict_df,
                        how='left',
                        left_on=config.drugs_header_name,     
                        right_on=config.cust_drug_names        
                    )
                    # Check for missing mappings
                    missing = merged[merged[config.std_drug_names].isna()]

                    if not missing.empty:
                        # Get original values for reporting
                        missing_report = merged.loc[missing.index, [config.drugs_header_name]]
                        missing_drugs[sheet] = list(missing_report[config.drugs_header_name].unique())
                except Exception as e:
                    error_manager.log_exception(f"Error validating unmapped drugs in sheet '{sheet}': {str(e)}")
                    raise
            
            # Error reporting
            if any(missing_drugs):
                error_msg = ["\n\n" + "=" * 50 + "\nUNMAPPED DRUGS FOUND IN DICTIONARY\n" + "=" * 50 + "\n"] 
        
                try:
                    wb = load_workbook(config.source_path_for_dictionary)
                    for sheet, drugs in missing_drugs.items():
                        ws = wb[sheet]
                        start_row = ws.max_row + 2
                        for i, value in enumerate(drugs):
                            ws.cell(row=start_row + i, column=1, value=value)
                        error_msg.append(f"➖ Sheet: {sheet} {len(drugs)} unmapped drugs.")
                    wb.save(config.source_path_for_dictionary)
                    wb.close()

                except Exception as e:
                    message = (
                    "\n🛠 HOW TO FIX:"
                    "\n• ❗Failed to write unmapped drugs to dictionary. You have to find the missing drugs yourself."
                    "\n• Add standard (mapped) names for customer-entered drug names in the dictionary."

                    "\n\n⚠️ Fix these issues in the dictionary file and try again."
                    )
                    error_msg.append(message)
                else:
                    message = (
                    "\n🛠 HOW TO FIX:"
                    "\n• Unmapped drug names already added the Customers_drug_names column."
                    "\n• Add standard (mapped) names for customer-entered drug names in the dictionary."

                    "\n\n⚠️ Fix these issues in the dictionary file and try again."
                    )
                    error_msg.append(message)

                finally:
                    if wb: wb.close()
                raise ExpectedCustomError('\n'.join(error_msg))
            error_manager.log_info("  No unmapped drugs found")
        except Exception:
            raise

    def replacing_drugs_to_standart(self):
        try:
            error_manager.log_info("Standardizing drug names...")
            for sheet, optovik_df in self.optoviks.items():
                # load dictionary
                dict_df = self.dm.drug_name_dict[sheet]
                
                # start replacing
                drug_map = dict_df.set_index(config.cust_drug_names)[config.std_drug_names]
                optovik_df[config.drugs_header_name] = optovik_df[config.drugs_header_name].map(drug_map)

                # store it
                self.optoviks[sheet] = optovik_df
            error_manager.log_info("  Drug names standardized successfully")
        except Exception:
            raise
        
    def load_budget_diff_drug_groups(self):
        try:
            error_manager.log_info("Loading budget difference and drug groups...")
            budget_dict = {}
            duplicate_errors = []  # Collect all duplicate errors

            # READ BUDGET DIFFERENCE
            self.budg_df = self.budg_df.rename(columns={
                self.budg_df.columns[config.budget_dif_drugs_col_index]: config.budg_dif_drugs,
                self.budg_df.columns[config.budget_dif_vtorichka_col_index]: config.budg_dif_vtorich,
                self.budg_df.columns[config.budget_dif_by_region_index]: config.budg_dif_by_reg
            })
            
            # Clean drug names and handle missing values
            self.budg_df[config.budg_dif_drugs] = (
                self.budg_df[config.budg_dif_drugs]
                .astype(str)
                .str.strip()
                .replace({'nan': pd.NA, '': pd.NA})
                .dropna()
            )
            
            # Check for duplicate drug names in budget data
            budget_duplicates = self.budg_df[self.budg_df.duplicated(subset=[config.budg_dif_drugs], keep=False)]
            if not budget_duplicates.empty:
                dup_drugs = budget_duplicates[config.budg_dif_drugs].unique()
                dup_list = "\n".join([f"- {drug}" for drug in dup_drugs])

                message_er = (
                    f"❌ Duplicate drugs found in budget difference file.\n"
                    f"({config.path_for_budget_difference.name}):\n{dup_list}"
                    )
                duplicate_errors.append(message_er)
                
            # 2. PROCESS DRUG GROUPS DATA
            
            # Melt and clean drug groups data
            drug_groups_df_melted = (
                pd.melt(self.drug_groups_df)
                .dropna()
                .rename(columns={'variable': 'Product_groups', 'value': 'Products_gr'})
            )
            
            # Clean drug names in melted data
            drug_groups_df_melted['Products_gr'] = (
                drug_groups_df_melted['Products_gr']
                .astype(str)
                .str.strip()
                .replace({'nan': pd.NA, '': pd.NA})
                .dropna()
            )
            
            # Check for duplicate drug names in drug groups
            drug_groups_duplicates = drug_groups_df_melted[drug_groups_df_melted.duplicated(subset=['Products_gr'], keep=False)]
            if not drug_groups_duplicates.empty:
                dup_drugs = drug_groups_duplicates['Products_gr'].unique()
                dup_list = "\n".join([f"- {drug}" for drug in dup_drugs])
                message_er2 = (
                    f"❌ Duplicate drugs found in drug groups file.\n"
                    f"({config.source_path_drug_groups.name}):\n{dup_list}"
                )
                duplicate_errors.append(message_er2)
            
            # Raise combined error if any duplicates found
            if duplicate_errors:
                formatted_errors = "\n\n".join(duplicate_errors)
                message = (
                    "\n\n" + "=" * 50 + "\nBUDGET DIFFERENCE OR DRUG GROUPS HAS DUBLICATES\n" + "=" * 50 + "\n\n"
                    f"{formatted_errors}"
                    "\n\n🛠 HOW TO FIX:"
                    "\n1. Open the Excel files listed above and locate the duplicate drug names."
                    "\n2. Ensure each drug appears **only once** in each file."
                    "\n   - In the **budget difference file**, remove or merge duplicate rows for the same drug."
                    "\n   - In the **drug groups file**, ensure each drug belongs to only one group."
                    "\n3. Check for subtle differences in spelling, casing, or spacing that may cause duplicates."
                    "\n4. Save the updated files."
                    "\n5. Re-run the script to continue processing."

                    "\n\n⚠️ The script cannot proceed until all missing drugs are correctly added."
                )
                raise ExpectedCustomError(message)
            
            # Add cleaned data to data manager
            # Create budget dictionary
            a = self.budg_df[config.budg_dif_drugs].tolist()
            b = self.budg_df[config.budg_dif_vtorich].tolist()
            c = self.budg_df[config.budg_dif_by_reg].tolist()
            budget_dict = {drug: (v, r) for drug, v, r in zip(a, b, c)}
            
            # Add to data manager
            self.dm.add_budget_dict(budget_dict)
            self.dm.add_drug_groups_df(self.drug_groups_df)
            self.dm.add_drug_groups_df_melted(drug_groups_df_melted.reset_index(drop=True))
            error_manager.log_info("  Budget difference and drug groups loaded successfully")
        except Exception:
            raise
    
    def validate_budget_diff_drug_groups(self):
        """Validate drug references between datasets"""
        try:
            error_manager.log_info("Validating budget difference and drug groups...")
            # Collect all unique drug names from optoviks
            all_optovik_drugs = set()
            for sheet, optovik_df in self.optoviks.items():
                # Clean and collect drug names
                drugs_series = optovik_df[config.drugs_header_name].dropna().astype(str).str.strip()
                all_optovik_drugs.update(drugs_series.unique())
            
            # Prepare reference drug lists
            budget_drugs = set(str(key).strip() for key in self.dm.budget_dict.keys())
            drug_group_drugs = set(self.dm.drug_groups_df_melted['Products_gr'].dropna().astype(str).str.strip().unique())
            
            # Identify missing drugs
            missing_in_budget = sorted(all_optovik_drugs - budget_drugs)
            missing_in_drug_groups = sorted(all_optovik_drugs - drug_group_drugs)
            
            # Export to text files if missing drugs exist
            if missing_in_budget or missing_in_drug_groups:
                # Write missing drugs to files
                try:
                    with open(config.path_for_unmatched_drugs_txt / "missing_in_budget.txt", 'w', encoding='utf-8') as f:
                        f.write("\n".join(missing_in_budget))
                    with open(config.path_for_unmatched_drugs_txt / "missing_in_drug_groups.txt", 'w', encoding='utf-8') as f:
                        f.write("\n".join(missing_in_drug_groups))
                except Exception as e:
                    error_manager.log_exception(f"Error writing missing drugs to files: {str(e)}")
                    raise
                
                # Prepare error message
                error_msg = ["\n\n" + "=" * 50 + "\nBUDGET DIFFERENCE OR DRUG GROUPS HAS MISSING DRUGS\n" + "=" * 50 + "\n"]
                
                if missing_in_budget:
                    message = (
                        f"❌ {len(missing_in_budget)} drugs missing in {config.path_for_budget_difference.name}."
                        f"\nThe list of missing drugs exported to this txt file: {config.path_for_unmatched_drugs_txt / 'missing_in_budget.txt\n'}"
                    )
                    error_msg.append(message)

                if missing_in_drug_groups:
                    message2 = (
                        f"❌ {len(missing_in_drug_groups)} drugs missing in {config.source_path_drug_groups.name}."
                        f"\nThe list of missing drugs exported to this txt file: {config.path_for_unmatched_drugs_txt / 'missing_in_drug_groups.txt'}"
                    )
                    error_msg.append(message2)

                error_msg.append("\n🛠 HOW TO FIX:"
                                "\n1. Open the text files listed above and review the missing drug names."
                                "\n2. Add the missing drugs to the correct reference Excel file:"
                                "\n   - Add drugs missing in `budget_difference.xlsx` to the **budget difference file**"
                                "\n   - Add drugs missing in `drug_groups.xlsx` to the **drug groups file**"
                                "\n3. Double-check for typos, inconsistent casing, or spacing issues."
                                "\n4. Save the updated files."
                                "\n5. Re-run the script to continue processing.")
                error_msg.append(f"\n⚠️ The script cannot proceed until all missing drugs are correctly added.")
                # Final full message string
                final_message = "\n".join(error_msg)
                raise ExpectedCustomError(final_message)
            error_manager.log_info("  Budget difference and drug groups... validated successfully")
        except Exception:
            raise

    def handling_reserve_columns(self):
        try:
            error_manager.log_info("Extracting reserve columns as a new optovik...")
            for v in config.reserve_column_values_list:
                a_df = pd.DataFrame()  # empty df to collect rows matching value 'v'
                
                for sheet, optovik_df in self.optoviks.items():

                    mask = optovik_df[config.reserve_header_name] == v
                    new_df = optovik_df[mask]

                    if not new_df.empty:
                        a_df = pd.concat([a_df, new_df], ignore_index=True)
                        # Remove rows from original
                        self.optoviks[sheet] = optovik_df[~mask].copy()

                if not a_df.empty:
                    a_df_clean = a_df.drop(columns=[config.reserve_header_name], errors='ignore').reset_index(drop=True)
                    self.dm.add_mapped_optovik(v, a_df_clean)

            # After filtering is complete, drop reserve column from remaining optoviks and add them to mapped
            for sheet, optovik_df in self.optoviks.items():
                cleaned_df = optovik_df.drop(columns=[config.reserve_header_name], errors='ignore').reset_index(drop=True)
                self.dm.add_mapped_optovik(sheet, cleaned_df)
            error_manager.log_info(f"  Processed {len(config.reserve_column_values_list)} reserve columns")
        except Exception as e:
            error_manager.log_exception(f"Error processing reserve columns: {str(e)}")
            raise

    def process_all(self):
        try:
            self.load_optoviks()
            self.dublicate_client_region_validation()
            self.load_dictionary()

            self.validate_dictionary()
            self.validate_unmapped_drugs()
            self.replacing_drugs_to_standart()

            self.load_budget_diff_drug_groups()
            self.validate_budget_diff_drug_groups()

            self.handling_reserve_columns()
        except Exception:
            raise

if __name__ == "__main__":
    try:
        data_manager = DataManager()
        validator = DrugMappingValidator(data_manager)    
        validator.process_all()
        error_manager.log_info("Validation processes completed successfully.")
    except ExpectedCustomError as e:
        error_manager.log_error(e)
    except Exception as e:
        error_manager.log_exception(e)

