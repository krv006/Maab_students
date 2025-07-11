import sys
from pathlib import Path
# Get current script's directory
current_dir = Path(__file__).resolve().parent
# Go up one level to project root
project_root = current_dir.parent
# Add project root to Python path
sys.path.append(str(project_root))

from common.config_handler import config
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

class OutlineAndFormulas():
    def __init__(self,last_row_indexes, list_for_outlining_columns, data_manager):
        self.dm = data_manager

        self.last_row_indexes = last_row_indexes
        self.list_for_outlining_columns = list_for_outlining_columns

        self.budget_dict = self.dm.budget_dict

        self.grouped_rows = []
        self.grouped_col_indexes = {}
        self.start_end_of_groups = {} 

    def apply_outline_to_customers(self,ws,sheet):

        # GROUPING ROWS
        start = False
        self.grouped_rows = []
        group_header_row = None  # the row where the formula should be written
        for row in range(1, ws.max_row + 1):
            col_a = ws.cell(row=row, column=1).value
            col_b = ws.cell(row=row, column=2).value

            if col_a == 0 and col_b in self.last_row_indexes[sheet]:
                start = True
                group_header_row = row  # this is where formulas will go
                # --
                self.grouped_rows.append(row) # save it for future use ---------------------------------------------
                # --
                group_start_row = row + 1  # start of actual hidden rows
                last_row_value = self.last_row_indexes[sheet][col_b]
                continue  # Start grouping from the next row

            if start:
                ws.row_dimensions[row].outlineLevel = 1
                ws.row_dimensions[row].hidden = True

            if start and col_a == last_row_value:   
                group_end_row = row  # end of the hidden rows group

                # Insert formulas at the header row (above the hidden rows)
                for col in range(5, ws.max_column + 1):  # skipping A and B
                    col_letter = get_column_letter(col)
                    formula = f"=SUM({col_letter}{group_start_row}:{col_letter}{group_end_row})"
                    ws.cell(row=group_header_row, column=col).value = formula

                # Reset for next group
                start = False
                group_header_row = None
        ws.sheet_properties.outlinePr.summaryBelow = False

        return ws

    def apply_outline_to_drugs(self,ws):
                # GROUPING COLUMNS
        group_col_numbers = []

        for col in range(1,ws.max_column+1):
            cell_value = ws.cell(row=1,column=col).value
            if cell_value in self.list_for_outlining_columns:
                group_col_numbers.append(col)
                # ----
                self.grouped_col_indexes[cell_value] = (col,col+1) # for future use
                # ---
        for ind in range(len(group_col_numbers)):
            start = group_col_numbers[ind] + 2

            if ind == len(group_col_numbers)-1:
                stop = ws.max_column - 1
            else:
                stop = group_col_numbers[ind + 1]

            for col in range(start, stop):
                col_letter = get_column_letter(col)

                ws.column_dimensions[col_letter].outline_level = 1
                ws.column_dimensions[col_letter].hidden = True
                ws.sheet_properties.outlinePr.summaryRight = False
            
        return ws

    # FORMULA PART
    def _applying_formulas_for_grouped_columns(self,ws, write_f_qty_col: bool, write_f_sales_col: bool, start_row: int, end_row: int):
                # PREPARING FOR FORMULAS
        self.start_end_of_groups = {}
        keys = list(self.grouped_col_indexes.keys())
        for i in range(len(keys) - 1):
            current_key = keys[i]
            next_key = keys[i + 1]
            start = self.grouped_col_indexes[current_key][1] + 1
            end = self.grouped_col_indexes[next_key][0] - 1
            self.start_end_of_groups[current_key] = (start, end)
        # Handle the last key ('Others')
        self.start_end_of_groups[keys[-1]] = (
            self.grouped_col_indexes[keys[-1]][1] + 1,
            ws.max_column - 2  # Assuming ws is defined earlier
        )

        # APPLY FORMULAS IN THE EACH GROUP
        for group, (start_col, end_col) in self.start_end_of_groups.items():

            qty_target_col, sales_target_col = self.grouped_col_indexes[group]

            # Separate quantity and sales columns in range
            quantity_cols = [col for i, col in enumerate(range(start_col, end_col + 1)) if i % 2 == 0]
            sales_cols = [col for i, col in enumerate(range(start_col, end_col + 1)) if i % 2 == 1]

            for row in range(start_row, end_row + 1):
                if row not in self.grouped_rows:
                    # Build SUM formulas
                    qty_formula = "=" + "+".join(f"{get_column_letter(col)}{row}" for col in quantity_cols)
                    sales_formula = "=" + "+".join(f"{get_column_letter(col)}{row}" for col in sales_cols)

                    if write_f_sales_col and write_f_qty_col:
                        ws.cell(row=row, column=sales_target_col).value = sales_formula
                        ws.cell(row=row, column=qty_target_col).value = qty_formula
                    elif write_f_qty_col:
                        ws.cell(row=row, column=qty_target_col).value = qty_formula
                    elif write_f_sales_col:
                        ws.cell(row=row, column=sales_target_col).value = sales_formula         

    def _applying_formulas_for_vertical_itogo_col(self,ws, start_row, end_row, skip_col: int):
                # VERTICAL ITOGO COLUMN FORMULA
        qty_col_itogo = ws.max_column - 1
        sales_col_itogo = ws.max_column
        quantity_cols_itogo = [t[0] for t in self.grouped_col_indexes.values()]
        sales_cols_itogo = [t[1] for t in self.grouped_col_indexes.values()]

        for row in range(start_row,end_row + 1):
            if row not in self.grouped_rows and row != skip_col:
                qty_formula_itogo = "=" + "+".join(f"{get_column_letter(col)}{row}" for col in quantity_cols_itogo)
                sales_formula_itogo = "=" + "+".join(f"{get_column_letter(col)}{row}" for col in sales_cols_itogo)

                ws.cell(row=row, column=qty_col_itogo).value = qty_formula_itogo
                ws.cell(row=row, column=sales_col_itogo).value = sales_formula_itogo

    def _is_percentage(self,value):
        return isinstance(value, float) and 0 < value <= 1

    def _apply_final_functions(self, ws, final_sum, final_sum_10, final_sum_rekl, final_sum_leks, sheet, total_sheet):
        
        # FINAL SUM ROW
        for col in range(5, ws.max_column + 1):
            formula_itogo_row = "=" + "+".join(f"{get_column_letter(col)}{row}" for row in self.grouped_rows)
            ws.cell(row=final_sum, column=col).value = formula_itogo_row
  
        # TOTAL SHEET Data preparation
        if total_sheet and sheet != config.vtorichka_sheet_name:
            self.dm.add_total_sheet_data(sheet,sheet)
            sales_cols_gr = [t[1] for t in self.grouped_col_indexes.values()]
            for col in range(5, ws.max_column +1):
                if col in sales_cols_gr:
                    link_total_sheet = f"='{sheet}'!{get_column_letter(col)}{final_sum}"
                    self.dm.add_total_sheet_data(sheet,link_total_sheet)

            sales_col_itogo = ws.max_column
            for row in [final_sum, final_sum_10, final_sum_rekl, final_sum_leks]:
                if row == final_sum:
                    self.dm.add_total_sheet_data(sheet,None)
                else:
                    link_total_sheet = f"='{sheet}'!{get_column_letter(sales_col_itogo)}{row}"
                    self.dm.add_total_sheet_data(sheet,link_total_sheet)
    

        # FRINAL 3 GROUPED COLUMNS SUM BY ROWS
        self._applying_formulas_for_grouped_columns(ws, write_f_qty_col=False, write_f_sales_col=True, start_row=final_sum_10, end_row=final_sum_leks )

        for group, (start_col, end_col) in self.start_end_of_groups.items():

            # FINAL SUM (Minus 10) and Final su for Leksiya
            quantity_cols = [col for i, col in enumerate(range(start_col, end_col + 1)) if i % 2 == 0]
            sales_cols = [col for i, col in enumerate(range(start_col, end_col + 1)) if i % 2 == 1]
            for col in sales_cols:
                ws.cell(row=final_sum_10, column=col).value = f"={get_column_letter(col)}{final_sum}*{config.final_sum_minus}%"
                ws.cell(row=final_sum_leks, column=col).value = f"={get_column_letter(col)}{final_sum_10}*{config.final_sum_leksiya}%"

            # FINAL SUM reklama
            for qty_col in quantity_cols:
                sales_col = qty_col+1
                drug = ws.cell(row=1, column=qty_col).value
                
                if drug not in self.budget_dict.keys():
                    ws.cell(row=final_sum_rekl, column=sales_col).value = "Drug name didn't match"
                    continue

                vtorich_val, regions_val = self.budget_dict[drug]

                # Determine multiplier based on sheet
                multiplier = vtorich_val if sheet == config.vtorichka_sheet_name else regions_val

                # Build formula
                if self._is_percentage(multiplier):
                    formula = f"={get_column_letter(sales_col)}{final_sum_10}*{multiplier}"
                else:
                    formula = f"={get_column_letter(qty_col)}{final_sum}*{multiplier}"
                ws.cell(row=final_sum_rekl, column=sales_col).value = formula

    def apply_excel_functions(self, ws, sheet , total_sheet=False):

        self.final_sum, final_sum_10, final_sum_rekl, self.final_sum_leks = ws.max_row+1, ws.max_row+2, ws.max_row+3, ws.max_row+4

        ws.cell(row=self.final_sum, column= 2).value= config.final_sum
        ws.cell(row=final_sum_10 , column= 2).value= config.final_sum_minus10
        ws.cell(row=final_sum_rekl , column= 2).value= config.final_sum_reklama
        ws.cell(row=self.final_sum_leks , column= 2).value= config.final_sum_leksiya_text

        start_row = 4
        self._applying_formulas_for_grouped_columns(ws, write_f_qty_col=True, write_f_sales_col = True, start_row=start_row, end_row=self.final_sum-1 )

        self._applying_formulas_for_vertical_itogo_col(ws, start_row=start_row, end_row=self.final_sum_leks, skip_col=self.final_sum)
        self._apply_final_functions(ws, self.final_sum, final_sum_10, final_sum_rekl, self.final_sum_leks, sheet, total_sheet)

        return ws


    def apply_formatting(self,ws):

        final_sum_row = self.final_sum

        header_color = PatternFill(
            start_color="F4ECC5", 
            end_color="F4ECC5", 
            fill_type="solid"
        )
        optovik_color = PatternFill(
            start_color="FFE4B5", 
            end_color="FFE4B5", 
            fill_type="solid"
        )
        default_border = Border(
            left=Side(style='thin', color='CCCC00'),
            right=Side(style='thin', color='CCCC00'),
            top=Side(style='thin', color='CCCC00'),
            bottom=Side(style='thin', color='CCCC00')
        )
        defult_font = Font(
            name="Arial",
            size=7,
            color="000000"
        )
        header_optovik_font = Font(
            name="Arial",
            size=8,
            bold=True,
            color="000000"
        )
        final_sums_font = Font(
            name="Arial",
            size=8,
            italic=True,
            color="000000"
        )

        # columns width

        ws.column_dimensions[get_column_letter(1)].width = 4
        ws.column_dimensions[get_column_letter(2)].width = 40
        ws.column_dimensions[get_column_letter(3)].width = 10
        ws.column_dimensions[get_column_letter(4)].width = 10

        for col in range(5,ws.max_column+1):
            if col % 2 == 0:
                ws.column_dimensions[get_column_letter(col)].width = 18
            else:
                ws.column_dimensions[get_column_letter(col)].width = 11
            

        # rows height
        ws.row_dimensions[1].height = 35
        ws.row_dimensions[2].height = 23
        
        # HEADER ROWS
        for row in ws.iter_rows(min_row=1, max_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                cell.border = default_border
                cell.fill = header_color
                cell.font = header_optovik_font

        # number formating cells
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:

                if cell.row in self.grouped_rows:
                    cell.fill = optovik_color
                    cell.border = default_border
                    if cell.column <= 4:
                        cell.font = header_optovik_font
                    elif cell.column > 4:
                        cell.font = defult_font
                        cell.number_format = '_(* #,##0.00_);_(* -#,##0.00_);_(* "-"??_);_(@_)'

                elif cell.column > 4 and cell.row < final_sum_row:
                    cell.number_format = '_(* #,##0.00_);_(* -#,##0.00_);_(* "-"??_);_(@_)'
                    cell.font = defult_font
                elif cell.row >= final_sum_row:
                    if cell.column > 4:
                        cell.number_format = '_(* #,##0.00_);_(* -#,##0.00_);_(* "-"??_);_(@_)'
                    cell.font = final_sums_font
                    cell.border = default_border
                    cell.fill = header_color

                elif cell.column <= 4 and cell.row < final_sum_row:
                    cell.font = defult_font

        # HIDE A ROW AND COLUMN
        ws.row_dimensions[3].hidden = True  # hides row 
        ws.column_dimensions['A'].hidden = True  # hides column

        # stuck header
        ws.freeze_panes = 'E3'

                

    def total_sheet_writer(self,ws):
        self.dm.add_total_sheet_data("Region","Region")
        self.dm.add_total_sheet_data("Region",self.list_for_outlining_columns, extend=True)
        for f_v in [config.final_sum, config.final_sum_minus10, config.final_sum_reklama, config.final_sum_leksiya_text]:
            self.dm.add_total_sheet_data("Region",f_v)

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # WRITING VALUES
        # writing header
        header_row = 2
        for col, value in enumerate(self.dm.total_sheet_data['Region'],start=2):
            cell = ws.cell(row=header_row, column=col, value=value)
            if value == config.final_sum:
                final_sum_col = col

            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.fill = yellow_fill

        del self.dm.total_sheet_data['Region']
        # insert all values
        for row, val_list in enumerate(self.dm.total_sheet_data.values(), start=3):
            for col, value in enumerate(val_list,start=2):
                ws.cell(row=row, column=col, value=value)

        # vertical itogo
        vertical_itogo_column = ws.max_column+1
        cell = ws.cell(row=header_row, column=vertical_itogo_column, value="Total for Reklama & for Leksiya")
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.fill = yellow_fill
        for row in range(3, ws.max_row+1):
            rek_col = get_column_letter(vertical_itogo_column-2)
            lek_col = get_column_letter(vertical_itogo_column-1)
            formula= f"={rek_col}{row} + {lek_col}{row}"
            ws.cell(row=row, column=vertical_itogo_column, value=formula)

        # horizontal itogo
        total_row = ws.max_row+1
        cell = ws.cell(row=total_row, column=2, value="Total")
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.fill = yellow_fill
        for col in range(3, ws.max_column + 1):
            col_letter = get_column_letter(col)
            
            formula = f"=SUM({col_letter}3:{col_letter}{total_row-1})"
            cell = ws.cell(row=total_row, column=col, value=formula)

            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.fill = yellow_fill

        # STYLING
        for row in range(header_row+1, total_row):
            cell = ws.cell(row=row, column=2)
            cell.alignment = Alignment(horizontal='left',vertical='center',wrap_text=True)
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.fill = green_fill

        # columns width
        ws.column_dimensions['A'].width = 4
        for col in range(2,ws.max_column+1):
            letter = get_column_letter(col)
            ws.column_dimensions[letter].width = 22

        # rows height
        ws.row_dimensions[2].height = 45
        
        # # alignment
        for row in ws.iter_rows(min_row=2, max_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)

        # # number formating cells

        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=3, max_col=ws.max_column):
            for cell in row:
                cell.number_format = '_(* #,##0.00_);_(* -#,##0.00_);_(* "-"??_);_(@_)'
                if cell.row != total_row:
                    cell.border = thin_border

                    if cell.column == final_sum_col:
                        cell.fill = yellow_fill
                        formula = f"=SUM(C{cell.row}:{get_column_letter(final_sum_col-1)}{cell.row})"
                        cell.value = formula
                    else:
                        cell.fill = green_fill
                        
