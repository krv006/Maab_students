import sys
from pathlib import Path
# Get current script's directory
current_dir = Path(__file__).resolve().parent
# Go up one level to project root
project_root = current_dir.parent
# Add project root to Python path
sys.path.append(str(project_root))

from common.config_handler import config
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from dashboard_automation.new_database_etl import SalesDataWarehouse
import pandas as pd
from datetime import datetime

class StagesProcesses():
    def __init__(self, data_manager):
        self.dm = data_manager
        sales_dw = SalesDataWarehouse()
        self.main_df = sales_dw.transform_data(self.dm.final_raw_optoviks, self.dm.drug_groups_df_melted, stage = True)

    def stage1_process(self,df):
        with pd.ExcelWriter("data/final/stage-1.xlsx", engine='openpyxl') as writer:
            for group_name, group in df.groupby("Product_groups"):
                pivoted_df = group.pivot_table(
                        index=["Region","Territory", "Product_groups"],
                        columns="Product",
                        values=["Quantity", "TotalSales"],
                        aggfunc='sum'
                    )
                
                columns = []
                for drug in pivoted_df.columns.get_level_values(1).unique():
                    columns.extend([("Quantity", drug), ( "TotalSales", drug)])
                pivoted_df = pivoted_df.reindex(columns=columns).reset_index()
                pivoted_df.columns = pivoted_df.columns.swaplevel(0, 1)

                pivoted_df.to_excel(writer, sheet_name=group_name, index=True)

                ws = writer.sheets[group_name]
                total_row = ws.max_row+1
                # STYLING
                # # alignment
                for row in ws.iter_rows(min_row=1, max_row=2):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                # # number formating cells
                for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=5, max_col=ws.max_column):
                    for cell in row:
                        cell.number_format = '_(* #,##0.00_);_(* -#,##0.00_);_(* "-"??_);_(@_)'
                                # columns width
                for col in range(2,ws.max_column+1):
                    letter = get_column_letter(col)
                    if col in [2,3,4]:
                        ws.column_dimensions[letter].width = 15
                    else:
                        if col % 2 == 1:
                            ws.column_dimensions[letter].width = 12
                        else:
                            ws.column_dimensions[letter].width = 18

                # rows height
                ws.row_dimensions[1].height = 45
                ws.row_dimensions[2].height = 35

            # ////////////////////////////////////////////////////
            all_teams_with_headers = []

            # Aggregate by both Product_groups and Product
            self.aggregated_data = df.groupby(['Product_groups', 'Product']).agg(
                total_quantity=('Quantity', 'sum'),
                total_sales=('TotalSales', 'sum')
            ).reset_index()

            # Group by Product_groups
            for group_name, group in self.aggregated_data.groupby("Product_groups"):
                group = group.drop(columns="Product_groups")
                group = group.reset_index(drop=True)
                group.index = group.index + 1
                group.index.name = "№"

                # Create empty row with only the group name in the first cell
                header_data = [group_name] + [''] * (group.shape[1] - 1)
                header_index = ['']  # Empty index column
                header_row = pd.DataFrame([header_data], columns=group.columns, index=header_index)

                # Append header and group
                all_teams_with_headers.append(pd.concat([header_row, group], axis=0))
            final_df = pd.concat(all_teams_with_headers)
            final_df.to_excel(writer, sheet_name="Проверка", index=True)

            self.styling_proverka(writer,"Проверка", self.aggregated_data, final_df, "92D050")

    def styling_proverka(self,writer,sheet_name, aggregated_data, final_df, color):
            # Now style each group header row
            sheet = writer.sheets[sheet_name]
            bold_font = Font(bold=True)
            row_index = 2  # Excel rows are 1-based, and row 1 is the column header
            gray_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            center_align = Alignment(horizontal='center', vertical='center')
            for group_name, group in aggregated_data.groupby("Product_groups"):
                # Apply bold font to this group's header row
                for col in range(1, len(group.columns) + 1):  # +1 for index, +1 for 1-based
                    cell = sheet.cell(row=row_index, column=col)
                    cell.font = bold_font
                    cell.fill = gray_fill
                    cell.alignment = center_align

                row_index += len(group) + 1  # 1 for the header row, len(group) for data rows


            # ////////////////////////////////////////////////////
            # Example: set column widths for first N columns
            num_columns = len(final_df.columns) + 1  # +1 if index is included

            for col_idx in range(1, num_columns + 1):  # 1-based indexing
                col_letter = get_column_letter(col_idx)
                if col_idx == 2:
                    sheet.column_dimensions[col_letter].width = 70  # Set width (adjust as needed)
                elif col_idx in range(3,sheet.max_column+1):
                    sheet.column_dimensions[col_letter].width = 22  # Set width (adjust as needed)

            # /////////////////////
            number_format = '_(* #,##0.00_);_(* -#,##0.00_);_(* "-"??_);_(@_)'
            target_col_indices = range(3,sheet.max_column+1)  # 1-based: Column C (3), D (4)

            for row in range(2, sheet.max_row + 1):
                for col_idx in target_col_indices:
                    col_letter = get_column_letter(col_idx)
                    sheet[f"{col_letter}{row}"].number_format = number_format

    def stage2_1_process(self,df):
        month_name = datetime.today().strftime('%B')
        list_stage2_total = []
        with pd.ExcelWriter("data/final/stage-2_Regionwise.xlsx", engine='openpyxl') as writer:
            # filter by region partf
            no_reserve_df = df[~df['Optovik'].isin(config.reserve_column_values_list)]
            only_reserve_df = df[df['Optovik'].isin(config.reserve_column_values_list)]

            def _stage_2_1_mini(main_df, group_by):
                for region_name, region_df in main_df.groupby(group_by):
                    aggregated_data = region_df.groupby(['Product', 'Product_groups']).agg(
                            total_quantity=('Quantity', 'sum')
                        ).reset_index()
                    renamed_df = aggregated_data.rename(columns={"total_quantity": region_name})
                    list_stage2_total.append(renamed_df)


                    all_teams_with_headers = []
                    # Group by Product_groups
                    for group_name, group in aggregated_data.groupby("Product_groups"):
                        group = group.drop(columns="Product_groups")
                        group = group.reset_index(drop=True)
                        group.index = group.index + 1
                        group.index.name = "№"

                        # Create empty row with only the group name in the first cell
                        header_data = [group_name] + [''] * (group.shape[1] - 1)
                        header_index = ['']  # Empty index column
                        header_row = pd.DataFrame([header_data], columns=group.columns, index=header_index)

                        # Append header and group
                        all_teams_with_headers.append(pd.concat([header_row, group], axis=0))

                    final_df = pd.concat(all_teams_with_headers)
                    
                    final_df.rename(columns={'total_quantity': month_name, 'Product':'Name of Products'}, inplace=True)
                    final_df.to_excel(writer,sheet_name=region_name, index=True)
                    self.styling_proverka(writer, region_name, aggregated_data, final_df, "E6B8B7")

            _stage_2_1_mini(no_reserve_df, 'Region')
            _stage_2_1_mini(only_reserve_df, 'Optovik')

            # ///////////TOTAL 2_1
        # with pd.ExcelWriter("data/stage-2_Productwise_total.xlsx", engine='openpyxl') as writer:
            dfs = [df.set_index(["Product", "Product_groups"]) for df in list_stage2_total]
            result = pd.concat(dfs, axis=1).reset_index()
            all_teams_with_headers2 = []
            # Group by Product_groups
            for group_name, group in result.groupby("Product_groups"):
                group = group.drop(columns="Product_groups")
                group = group.reset_index(drop=True)
                group.index = group.index + 1
                group.index.name = "№"

                # Create empty row with only the group name in the first cell
                header_data = [group_name] + [''] * (group.shape[1] - 1)
                header_index = ['']  # Empty index column
                header_row = pd.DataFrame([header_data], columns=group.columns, index=header_index)

                # Append header and group
                all_teams_with_headers2.append(pd.concat([header_row, group], axis=0))

            final_df2 = pd.concat(all_teams_with_headers2)
            final_df2.rename(columns={'Product':'Name of Products'}, inplace=True)
            final_df2.to_excel(writer,sheet_name="Total", index=True)

            # STIYLE IT/////////////////////
            ws = writer.sheets["Total"]
            product_groups = df['Product_groups'].dropna().unique().tolist()

            gr_fill_style = PatternFill(fill_type='solid', start_color='D99795', end_color='D99795') 
            hd_fill_style = PatternFill(fill_type='solid', start_color='CCFFCC', end_color='CCFFCC')
            gr_font_style = Font(bold=True)
            gr_alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=True

            )
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    if cell.value in product_groups:
                        row_number = cell.row
                        for c in ws[row_number]:
                            c.fill = gr_fill_style
                            c.font = gr_font_style
                            c.alignment = gr_alignment
                    elif cell.row == 1:
                        cell.fill = hd_fill_style
                        cell.font = gr_font_style
                        cell.alignment = gr_alignment
                    # Apply number format to all except headers and excluded columns
                    elif isinstance(cell.value, (int, float)) and cell.row > 1 and cell.column not in [1,2]:
                        cell.number_format = '_(* #,##0.00_);_(* -#,##0.00_);_(* "-"??_);_(@_)'
            
            ws.row_dimensions[1].height = 25


            for col_idx in range(1, ws.max_column + 1):  # 1-based indexing
                col_letter = get_column_letter(col_idx)
                if col_idx == 2:
                    ws.column_dimensions[col_letter].width = 70  # Set width (adjust as needed)
                elif col_idx in range(3,ws.max_column+1):
                    ws.column_dimensions[col_letter].width = 12
          
    def stage2_2_process(self,df):

        with pd.ExcelWriter("data/final/stage-2_Total_sales.xlsx", engine='openpyxl') as writer:

            aggregated_data = df.groupby(['Product', 'Product_groups']).agg(
                total_sales=('TotalSales', 'sum')
            ).reset_index()
            renamed_df = aggregated_data.rename(columns={"total_sales": "Total Sales", "Product": "Name of Products", "Product_groups":"Group"})
            renamed_df.to_excel(writer, sheet_name="Sheet_", index=True)
            ws = writer.sheets["Sheet_"]

            total_row = ws.max_row+1
            # STYLING
            # # alignment
            for row in ws.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            # # number formating cells
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=ws.max_column):
                for cell in row:
                    cell.number_format = '_(* #,##0.00_);_(* -#,##0.00_);_(* "-"??_);_(@_)'

            # columns width
            for col_idx in range(1, ws.max_column + 1):  # 1-based indexing
                col_letter = get_column_letter(col_idx)
                if col_idx == 2:
                    ws.column_dimensions[col_letter].width = 70  # Set width (adjust as needed)
                elif col_idx in range(3,ws.max_column+1):
                    ws.column_dimensions[col_letter].width = 22

            # rows height
            ws.row_dimensions[1].height = 45

    def run_stage(self):
        self.stage1_process(self.main_df)
        self.stage2_1_process(self.main_df)
        self.stage2_2_process(self.main_df)







