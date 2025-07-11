import sys
from pathlib import Path

# Get current script's directory
current_dir = Path(__file__).resolve().parent
# Go up one level to project root
project_root = current_dir.parent
# Add project root to Python path
sys.path.append(str(project_root))


from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from common.config_handler import config
from common.error_handler import error_manager, ExpectedCustomError

class Processing_1c_source:
    def __init__(self):
        self.source_path = config.source_path_for_1c
        self.output_path_for_pivoted = config.path_for_1c_pivoted

        self._check_required_file()

        # Attributes
        self.wb = None
        self.ws = None
        self.headers = []
        self.last_row = None

        self.rows_to_delete = []
        self.cols_to_delete = []

        # Creating new workbook to write
        self.new_wb = Workbook()
        self.new_ws = self.new_wb.active
        self.new_ws.title = config.sheet_name_1c

    def _check_required_file(self):
        required_files = {
            "raw_1c_file": config.source_path_for_1c
        }

        missing_files = []
        for name, path in required_files.items():
            if not Path(path).exists():
                missing_files.append((name, path))

        if missing_files:
            error_lines = ["\n\n" + "=" * 50 + "\nA FILE is MISSING\n" + "=" * 50 + "\n"]
            error_lines.append("\n❌ The following required files are missing:\n")
            for name, path in missing_files:
                error_lines.append(f"• {name} — Missing file: {Path(path).name}\n  📁 Expected at: {path}\n")

            error_lines.append("\n📌 HOW TO FIX:\n")
            error_lines.append("1. Make sure the file listed above exists at its expected path.\n")
            error_lines.append("2. If missing, export or copy the correct version into the folder.\n")
            error_lines.append("3. Make sure none of the file is open in Excel or locked by another program.\n")
            error_lines.append("4. If the file path is wrong in config, update them accordingly.\n")
            error_lines.append("\n⚠️ The script cannot continue unless required file is present in expected location.")

            raise ExpectedCustomError("".join(error_lines))

    def loading_source_workbook(self):
        try:
            self.wb = load_workbook(self.source_path)
            self.ws = self.wb.active
        except FileNotFoundError:
            raise FileNotFoundError(f"Source workbook not found at '{self.source_path}'")
        except Exception as e:
            raise RuntimeError(f"Error loading source workbook: {e}")

    def detect_headers(self):
        for row in range(1,self.ws.max_row + 1):
            lvl = self.ws.row_dimensions[row].outline_level
            if lvl == 1: 
                for row2 in range(row+1,self.ws.max_row + 1):
                    nxt_lvl = self.ws.row_dimensions[row2].outline_level
                    if nxt_lvl < lvl:
                        break
                    elif nxt_lvl > lvl:
                        self.headers =[row-4,row-3,row-2]
                        return
        if not self.headers:
            raise ValueError("Failed to detect headers in source file")
        
    
    def finding_table_boundries(self):
        # finding the last row of the table
        for row in range(self.ws.max_row,0,-1):
            if self.ws.row_dimensions[row].outline_level > 0:
                self.last_row = row
                break

        if not self.last_row:
            raise ValueError("No data rows found in source worksheet")

        # delete the rows out of table range
        self.rows_to_delete=[row for row in range(1,min(self.headers))]
        self.rows_to_delete.append(max(self.headers))
        for row in range(self.last_row+1,self.ws.max_row+1):
            self.rows_to_delete.append(row)

    def identify_columns_to_remove(self):
        # iterate and find must delete columns
        for col in range(1, self.ws.max_column + 1):
            if col in [1,2,3]: #skip first 3 columns
                continue
            if self.ws.column_dimensions[get_column_letter(col)].outline_level in [0,2]:
                self.cols_to_delete.append(col)

    def process_rows(self):
        for row_ind,row  in enumerate(self.ws.iter_rows(values_only=True),1): 
            row = list(row)
            # Renaming and cleaning some column names
            if row_ind == self.headers[1]:
                row[0] = config.client_header_name
                row[1] = config.region_header_name
                row[2] = config.territory_header_name
            elif row_ind == self.headers[0]:
                row[0] = None
                row[1] = None
                row[2] = None

            lvl = self.ws.row_dimensions[row_ind].outline_level
            next_lvl = self.ws.row_dimensions[row_ind+1].outline_level

            # Handling region names and must delete rows
            if lvl == 0 and next_lvl == 1 and row_ind not in self.rows_to_delete:
                region = row[0]
                self.rows_to_delete.append(row_ind)
            elif lvl == 1 and row_ind not in self.rows_to_delete:
                subregion = row[0]
                self.rows_to_delete.append(row_ind)
            elif lvl == 2 and row_ind not in self.rows_to_delete:
                row[1]=region
                row[2]=subregion
            # self.row_values.append({'row_ind':row_ind,'outline_level':lvl,'row_values':row})

            # WRITING ROWS
            if row_ind == self.headers[1] + 1:
                self.new_ws.append([])# Adds an empty row after header
            if row_ind not in self.rows_to_delete:
                filtered_row = [value for col_num, value in enumerate(row,1) if col_num not in self.cols_to_delete]
                self.new_ws.append(filtered_row)

    def apply_formatting(self):
        # merging header
        for cell in range(4,self.new_ws.max_column+1,2):
            self.new_ws.merge_cells(start_row=1,end_row=1,start_column=cell,end_column=cell+1)

        # number formating cells
        for row in self.new_ws.iter_rows(min_row=3, max_row=self.new_ws.max_row, min_col=4, max_col=self.new_ws.max_column):
            for cell in row:
                cell.number_format = '_(* #,##0.00_);_(* -#,##0.00_);_(* "-"??_);_(@_)'

    def save_workbook(self):
        try:
            self.new_wb.save(self.output_path_for_pivoted)
            error_manager.log_info(f"  A temporary file saved as: {self.output_path_for_pivoted.name}")
        except PermissionError:
            raise PermissionError(f"Permission denied: '{self.output_path_for_pivoted}' is open or write-protected")
        except Exception as e:
            raise RuntimeError(f"Error saving workbook: {e}")
        finally:
            self.new_wb.close()

    def process(self):
        try:
            self.loading_source_workbook()
            self.detect_headers()
            self.finding_table_boundries()
            self.identify_columns_to_remove()
            self.process_rows()
            self.apply_formatting()
            self.save_workbook()
            
        except Exception as e:
            raise  # Re-raise the exception to be caught in the main block

if __name__ == '__main__':
    print("cannot run from here")
